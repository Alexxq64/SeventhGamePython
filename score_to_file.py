from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
import time
from openpyxl import load_workbook


def extract_data_from_page(driver):
    """
    Извлекает информацию о счёте всех геймов с текущей страницы (point-by-point),
    а также информацию о подающем игроке в первом гейме первого сета.

    Аргументы:
        driver: WebDriver.

    Возвращает:
        data (list): Список всех данных о счёте геймов (без лишних символов).
        server_info (str): Информация о подающем игроке в первом гейме.
    """
    game_data = []
    server_info = "Неизвестно"

    try:
        for i in range(2, 27, 2):
            try:
                score_selector = f"#detail > div.matchHistoryRowWrapper > div:nth-child({i}) > div.matchHistoryRow__scoreBox"
                score_element = driver.find_element(By.CSS_SELECTOR, score_selector)
                game_data.append(score_element.text.replace("\n", ""))

                if i == 2:
                    serve_selector_left = f"#detail > div.matchHistoryRowWrapper > div:nth-child({i}) > div.matchHistoryRow__servis.matchHistoryRow__home > div > svg"
                    serve_selector_right = f"#detail > div.matchHistoryRowWrapper > div:nth-child({i}) > div.matchHistoryRow__servis.matchHistoryRow__away > div > svg"
                    try:
                        if driver.find_element(By.CSS_SELECTOR, serve_selector_left):
                            server_info = "Игрок 1 подает"
                    except:
                        try:
                            if driver.find_element(By.CSS_SELECTOR, serve_selector_right):
                                server_info = "Игрок 2 подает"
                        except:
                            pass
            except Exception:
                break

        return game_data, server_info
    except Exception as e:
        print(f"Ошибка при извлечении данных: {e}")
        return [], "Неизвестно"


def convert_score_to_letters(scores):
    """
    Преобразует список счётов в формате ['X-Y', ...] в строку букв,
    определяя прирост выигрышей (A или B).

    Аргументы:
        scores (list): Список строк с результатами, например ['1-0', '2-0', '2-1', ...].

    Возвращает:
        str: Строка из букв 'A' и 'B', например 'AABAB...'.
    """
    letters = []
    previous_score = [0, 0]

    for score in scores:
        current_score = list(map(int, score.split('-')))
        delta_player1 = current_score[0] - previous_score[0]
        delta_player2 = current_score[1] - previous_score[1]

        if delta_player1 > delta_player2:
            letters.append('A')
        elif delta_player2 > delta_player1:
            letters.append('B')

        previous_score = current_score

    return ''.join(letters)


def switch_tabs_and_collect_data(driver):
    """
    Последовательно переключается между вкладками и собирает информацию с каждой.

    Аргументы:
        driver: WebDriver.

    Возвращает:
        all_data (dict): Словарь с данными по каждой вкладке.
    """
    all_data = {}
    try:
        tab_buttons = driver.find_elements(
            By.CSS_SELECTOR,
            "#detail > div.subFilterOver.subFilterOver--indent > div > a > button"
        )

        print(f"Найдено вкладок: {len(tab_buttons)}")

        for i, button in enumerate(tab_buttons):
            try:
                print(f"Переход на вкладку {i}...")
                button.click()
                time.sleep(2)

                data, server_info = extract_data_from_page(driver)
                data_letters = convert_score_to_letters(data)

                all_data[f"point-by-point/{i}"] = data_letters
                if i == 0:
                    all_data["server_info"] = server_info

            except Exception as e:
                print(f"Ошибка при переходе на вкладку {i}: {e}")
    except Exception as e:
        print(f"Ошибка при работе с вкладками: {e}")

    return all_data


def write_to_excel(file_path, parsed_data, match_url):
    """
    Записывает собранные данные в Excel-файл.

    Аргументы:
        file_path (str): Путь к Excel-файлу.
        parsed_data (dict): Словарь с данными, включая информацию о подающем и сетах.
        match_url (str): URL матча.
    """
    try:
        workbook = load_workbook(file_path)
        sheet = workbook["Sets"]

        # Ищем последнюю строку с данными в колонке 12
        last_row = sheet.max_row
        while sheet.cell(row=last_row, column=12).value is None and last_row > 1:
            last_row -= 1

        print(f"Последняя заполненная строка (с данными в колонке 12): {last_row}")

        server_info = parsed_data.get("server_info", "Неизвестно")
        server_value = "A" if server_info == "Игрок 1 подает" else "B" if server_info == "Игрок 2 подает" else ""
        target_row = last_row + 1

        # Записываем ссылку на матч в 4-й столбец

        for set_index, (key, score) in enumerate(parsed_data.items()):
            sheet.cell(row=target_row, column=4).value = match_url
            if key == "server_info":
                continue

            if set_index == 0 and server_value:
                sheet.cell(row=target_row, column=9).value = server_value

            for i, letter in enumerate(score, start=12):
                sheet.cell(row=target_row, column=i).value = letter
            target_row = target_row + 1  # Следующая пустая строка

        workbook.save(file_path)
        print(f"Данные успешно записаны в файл: {file_path}")

    except Exception as e:
        print(f"Ошибка при записи в файл: {e}")


def process_all_match_links(excel_file_path):
    """
    Читает ссылки на матчи с листа MatchLinks и обрабатывает их.

    Аргументы:
        excel_file_path (str): Путь к Excel-файлу.
    """
    workbook = load_workbook(excel_file_path)
    sheet = workbook["MatchLinks"]

    # Получаем все ссылки из первого столбца
    match_links = [sheet.cell(row=i, column=1).value for i in range(2, sheet.max_row + 1) if sheet.cell(row=i, column=1).value]

    for match_url in match_links:
        print(f"Обработка матча: {match_url}")
        process_match_page(match_url, excel_file_path)


def process_match_page(match_url, excel_file_path):
    """
    Переходит на страницу матча, собирает данные со всех вкладок point-by-point и записывает в Excel.

    Аргументы:
        match_url (str): URL страницы матча.
        excel_file_path (str): Путь к Excel-файлу.
    """
    chrome_options = Options()
    chrome_options.add_argument("--headless")
    chrome_options.add_argument("--disable-gpu")
    chrome_options.add_argument("--no-sandbox")
    chrome_service = Service("C:\\Users\\User\\Desktop\\Настройки\\chromedriver.exe")

    driver = webdriver.Chrome(service=chrome_service, options=chrome_options)

    try:
        driver.get(match_url)
        time.sleep(3)

        try:
            button_selector = "#detail > div.filterOver.filterOver--indent > div > a:nth-child(3) > button"
            button = driver.find_element(By.CSS_SELECTOR, button_selector)
            button.click()
            time.sleep(2)
        except Exception as e:
            print(f"Не удалось нажать кнопку 'point-by-point': {e}")
            return

        all_data = switch_tabs_and_collect_data(driver)

        if "server_info" in all_data:
            print(f"Информация о подающем игроке: {all_data['server_info']}")
        print("Собранные данные со всех вкладок:")
        for tab, data in all_data.items():
            if tab != "server_info":
                print(f"{tab}: {data}")

        write_to_excel(excel_file_path, all_data, match_url)

    finally:
        driver.quit()


if __name__ == "__main__":
    excel_file_path = "D:\\Hobby\\SeventhGamePython\\7thGame.xlsx"
    process_all_match_links(excel_file_path)
