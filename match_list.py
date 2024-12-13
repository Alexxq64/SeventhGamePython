from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
import os
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import time


def save_match_links_selenium(tournament_url, output_to_file=True):
    """
    Извлекает ссылки на матчи с использованием Selenium и сохраняет их в Excel файл или выводит в консоль.

    Параметры:
        tournament_url (str): URL страницы турнира.
        output_to_file (bool): Если True, сохраняет в файл; если False, выводит в консоль.
    """
    # Настройка Selenium WebDriver
    chrome_options = Options()
    chrome_options.add_argument("--headless")  # Открытие браузера в фоновом режиме
    chrome_options.add_argument("--disable-gpu")
    chrome_options.add_argument("--no-sandbox")
    chrome_service = Service("C:\\Users\\User\\Desktop\\Настройки\\chromedriver.exe")  # Укажите путь к chromedriver

    driver = webdriver.Chrome(service=chrome_service, options=chrome_options)

    try:
        # Открываем страницу турнира
        driver.get(tournament_url)
        time.sleep(3)  # Ждём загрузки страницы

        # Находим все элементы ссылок на матчи
        match_links = []
        elements = driver.find_elements(By.CSS_SELECTOR, "a[href*='game-summary']")

        for element in elements:
            match_url = element.get_attribute("href")
            if match_url and match_url not in match_links:  # Уникальные ссылки
                match_links.append(match_url)

        if output_to_file:
            # Путь к файлу Excel
            file_path = "D:\\Hobby\\SeventhGamePython\\7thGame.xlsx"

            try:
                # Загружаем файл Excel или создаем новый
                if os.path.exists(file_path):
                    wb = load_workbook(file_path)
                else:
                    wb = Workbook()

                # Создаем новый лист, если его нет
                if "MatchLinks" not in wb.sheetnames:
                    ws = wb.create_sheet("MatchLinks")
                    ws.cell(1, 1, "Ссылка на матч")  # Заголовок
                else:
                    ws = wb["MatchLinks"]

                # Начинаем запись с первой свободной строки
                start_row = ws.max_row + 1
                for link in match_links:
                    ws.cell(start_row, 1, link)
                    ws.cell(start_row, 1).hyperlink = link
                    start_row += 1

                # Автоподгонка ширины столбцов
                column = get_column_letter(1)
                max_length = max(len(str(ws[f"{column}{r}"].value or "")) for r in range(1, ws.max_row + 1))
                ws.column_dimensions[column].width = max_length + 2

                # Сохраняем файл Excel
                wb.save(file_path)
                print(f"Ссылки на матчи сохранены в Excel файл по пути: {file_path}")

            except PermissionError:
                print(f"Ошибка! Файл '{file_path}' в данный момент открыт. Закройте файл и попробуйте снова.")
            except Exception as e:
                print(f"Произошла ошибка: {e}")

        else:
            # Выводим в консоль
            print("Ссылки на матчи:")
            for link in match_links:
                print(link)

    finally:
        # Закрываем браузер
        driver.quit()


if __name__ == "__main__":
    tournament_url = "https://www.livesport.com/tennis/atp-singles/antwerp/results/"
    # Выберите, куда выводить результаты: True для файла, False для консоли
    save_match_links_selenium(tournament_url, output_to_file=True)
