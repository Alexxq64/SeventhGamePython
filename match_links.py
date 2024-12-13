from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import os

def extract_point_links(tournament_url):
    """
    Извлекает ссылки на вкладки "point-by-point" для матчей и сохраняет их в Excel файл.

    Параметры:
        tournament_url (str): URL страницы турнира.
    """
    # Настраиваем Selenium
    chrome_options = Options()
    chrome_options.add_argument("--headless")  # Открытие в фоновом режиме
    chrome_service = Service("C:\\Users\\User\\Desktop\\Настройки\\chromedriver.exe")

    # Инициализация браузера
    driver = webdriver.Chrome(service=chrome_service, options=chrome_options)
    wait = WebDriverWait(driver, 10)  # Ожидание до 10 секунд

    # Путь к Excel файлу
    file_path = "D:\\Hobby\\SeventhGamePython\\7thGame.xlsx"

    try:
        # Открываем страницу турнира
        driver.get(tournament_url)

        # Ожидаем загрузки хотя бы одного элемента
        try:
            wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "div.event__match")))
        except:
            print("Матчи не найдены на странице.")
            return

        # Загружаем файл Excel или создаём новый
        wb = load_workbook(file_path) if os.path.exists(file_path) else Workbook()
        ws_set_links = wb.create_sheet("SetLinks") if "SetLinks" not in wb.sheetnames else wb["SetLinks"]

        # Заголовки
        ws_set_links.cell(1, 1, "Ссылка на матч")
        ws_set_links.cell(1, 2, "Ссылка на вкладку")

        set_link_row = 2

        # Ищем матчи и обрабатываем
        while True:
            try:
                # Повторно находим список матчей
                matches = driver.find_elements(By.CSS_SELECTOR, "div.event__match")

                # Проверка: есть ли матчи на странице
                if not matches:
                    print("Список матчей пуст.")
                    break

                # Обрабатываем каждый матч
                for match in matches:
                    try:
                        # Извлекаем ссылку на матч
                        link_element = match.find_element(By.CSS_SELECTOR, "a.eventRowLink")
                        match_url = link_element.get_attribute("href")

                        # Переходим по ссылке на матч
                        for set_number in range(5):  # Проверяем вкладки от 0 до 4
                            point_by_point_url = f"{match_url}point-by-point/{set_number}"
                            driver.get(point_by_point_url)

                            # Проверяем, загрузилась ли нужная страница
                            try:
                                wait.until(EC.url_contains("point-by-point"))
                                # Сохраняем рабочую ссылку
                                ws_set_links.cell(set_link_row, 1, match_url)
                                ws_set_links.cell(set_link_row, 2, point_by_point_url)
                                set_link_row += 1
                            except:
                                break  # Если вкладка недоступна, выходим из цикла

                    except Exception as e:
                        print(f"Ошибка обработки матча: {e}")
                        continue  # Переходим к следующему матчу

                break  # Завершаем основной цикл после обработки матчей

            except Exception as e:
                print(f"Ошибка при обновлении списка матчей: {e}")
                break

        # Автоподгонка ширины столбцов
        for col in range(1, ws_set_links.max_column + 1):
            column = get_column_letter(col)
            max_length = max(len(str(ws_set_links[f'{column}{row}'].value or "")) for row in range(1, ws_set_links.max_row + 1))
            ws_set_links.column_dimensions[column].width = max_length + 2

        # Сохраняем файл Excel
        wb.save(file_path)

        # Выводим сообщение об успехе
        print(f"Данные о вкладках 'point-by-point' сохранены в Excel файл по пути: {file_path}")

    finally:
        # Закрываем браузер
        driver.quit()

# Основная функция
def main():
    tournament_url = "https://www.livesport.com/tennis/atp-singles/adelaide/results/"
    extract_point_links(tournament_url)

if __name__ == "__main__":
    main()
