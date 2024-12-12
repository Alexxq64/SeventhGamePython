import os
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import requests
from bs4 import BeautifulSoup

def save_tournaments_to_excel(base_url, file_path="D:\\Hobby\\SeventhGamePython\\7thGame.xlsx"):
    """
    Сохраняет список турниров в Excel файл, основываясь на переданном URL.

    Параметры:
        base_url (str): URL главной страницы с турнирами.
        file_path (str): Путь для сохранения Excel файла. По умолчанию "D:\\Hobby\\SeventhGamePython\\7thGame.xlsx".
    """
    # Отправляем GET запрос к сайту
    response = requests.get(base_url)

    # Проверяем успешность запроса
    if response.status_code != 200:
        print("Ошибка доступа к сайту.")
        return

    # Используем BeautifulSoup для парсинга HTML
    soup = BeautifulSoup(response.text, "html.parser")

    # Находим все ссылки на турниры
    tournament_links = soup.find_all("a", href=True)

    # Проверяем, доступен ли файл для записи
    if os.access(file_path, os.W_OK):
        try:
            # Создаем новый Excel файл
            wb = Workbook()
            ws = wb.active
            ws.title = "Tournaments"

            # Заголовки
            ws.cell(1, 1, "Название турнира")
            ws.cell(1, 2, "URL")
            ws.cell(1, 3, "URL результатов")

            row = 2

            # Обрабатываем все ссылки на турниры
            for link in tournament_links:
                href = link['href']
                
                # Проверяем, содержит ли ссылка "/tennis/atp-singles/" и имеет ли она продолжение
                if "/tennis/atp-singles/" in href and len(href) > len("/tennis/atp-singles/"):
                    # Формируем полный URL для турнира
                    if href.startswith("/"):
                        tournament_url = "https://www.livesport.com" + href
                    elif href.startswith("about:"):
                        tournament_url = href.replace("about:", "https://www.livesport.com")
                    else:
                        tournament_url = href
                    
                    # Формируем URL для результатов турнира
                    tournament_results_url = tournament_url + "results/"

                    # Получаем название турнира (текст ссылки)
                    tournament_name = link.get_text()

                    # Записываем данные в Excel
                    ws.cell(row, 1, tournament_name)  
                    ws.cell(row, 2, tournament_url)  
                    ws.cell(row, 3, tournament_results_url)  

                    # Создаем гиперссылки в Excel
                    ws.cell(row, 2).hyperlink = tournament_url  
                    ws.cell(row, 3).hyperlink = tournament_results_url  

                    # Переходим к следующей строке
                    row += 1

            # Автоподгонка ширины столбцов
            for col in range(1, 4):
                max_length = 0
                column = get_column_letter(col)
                for row in range(1, wb.active.max_row + 1):
                    cell_value = str(ws[f'{column}{row}'].value)
                    max_length = max(max_length, len(cell_value))
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column].width = adjusted_width

            # Сохраняем файл Excel
            wb.save(file_path)

            # Выводим сообщение об успехе
            print(f"Данные о турнирах сохранены в Excel по пути: {file_path}")

        except PermissionError:
            print(f"Ошибка! Файл '{file_path}' в данный момент открыт. Закройте файл и попробуйте снова.")

    else:
        print(f"Ошибка! Нет прав на запись в файл '{file_path}'. Убедитесь, что файл не открыт и доступен для записи.")



def save_match_links_to_excel(tournament_url):
    """
    Извлекает ссылки на матчи с заданной страницы турнира и сохраняет их в Excel файл.

    Параметры:
        tournament_url (str): URL страницы турнира.
    """
    # Отправляем GET запрос к странице турнира
    response = requests.get(tournament_url)

    # Проверяем успешность запроса
    if response.status_code != 200:
        print("Ошибка доступа к турниру.")
        return

    # Используем BeautifulSoup для парсинга HTML
    soup = BeautifulSoup(response.text, "html.parser")

    # Находим все элементы матчей с классом "event__match"
    match_elements = soup.find_all("div", class_="event__match")

    # Открываем существующий Excel файл или создаем новый
    file_path = "D:\\Hobby\\SeventhGamePython\\7thGame.xlsx"
    try:
        # Загружаем файл Excel
        wb = load_workbook(file_path) if os.path.exists(file_path) else Workbook()
        ws = wb.create_sheet("MatchLinks") if "MatchLinks" not in wb.sheetnames else wb["MatchLinks"]

        # Заголовки
        ws.cell(1, 1, "Ссылка на матч")

        row = 2

        # Обрабатываем каждый элемент матча
        for match in match_elements:
            # Находим тег <a> со ссылкой
            link_tag = match.find("a", href=True, class_="eventRowLink")
            if link_tag:
                match_url = link_tag['href']

                # Проверяем, является ли ссылка полной или относительной
                if match_url.startswith("/"):
                    match_url = "https://www.livesport.com" + match_url

                # Записываем ссылку в Excel
                ws.cell(row, 1, match_url)
                ws.cell(row, 1).hyperlink = match_url

                # Переходим к следующей строке
                row += 1

        # Автоподгонка ширины столбцов
        column = get_column_letter(1)
        max_length = max(len(str(ws[f'{column}{r}'].value)) for r in range(1, ws.max_row + 1))
        ws.column_dimensions[column].width = max_length + 2

        # Сохраняем файл Excel
        wb.save(file_path)

        # Выводим сообщение об успехе
        print(f"Ссылки на матчи сохранены в Excel файл по пути: {file_path}")

    except PermissionError:
        print(f"Ошибка! Файл '{file_path}' в данный момент открыт. Закройте файл и попробуйте снова.")

# Основная функция
def main():
    # Вызов функции с конкретным URL для турниров ATP
    save_tournaments_to_excel("https://www.livesport.com/tennis/calendar/atp/")
    
    # Вызов функции с конкретным URL для турниров WTA и другой путь для сохранения
    save_tournaments_to_excel("https://www.livesport.com/tennis/calendar/wta/", "D:\\Hobby\\SeventhGamePython\\7thGameWTA.xlsx")

    tournament_url = "https://www.livesport.com/tennis/atp-singles/adelaide/results/"
    save_match_links_to_excel(tournament_url)


# Блок для запуска main() только если файл запускается напрямую
if __name__ == "__main__":
    main()