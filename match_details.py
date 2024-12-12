from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import os
import time


def extract_matches_with_selenium(tournament_url):
    """
    Использует Selenium для извлечения информации о матчах с сайта Livesport и сохранения её в Excel.

    Параметры:
        tournament_url (str): URL страницы турнира.
    """
    # Настроим Selenium
    chrome_options = Options()
    chrome_options.add_argument("--headless")  # Открытие в фоновом режиме
    chrome_service = Service("C:\\Users\\User\\Desktop\\Настройки\\chromedriver.exe")

    # Инициализация браузера
    driver = webdriver.Chrome(service=chrome_service, options=chrome_options)

    # Открываем существующий Excel файл или создаем новый
    file_path = "D:\\Hobby\\SeventhGamePython\\7thGame.xlsx"
    wb = load_workbook(file_path) if os.path.exists(file_path) else Workbook()
    ws = wb.create_sheet("MatchDetails") if "MatchDetails" not in wb.sheetnames else wb["MatchDetails"]

    # Заголовки
    ws.cell(1, 1, "Название турнира")
    ws.cell(1, 2, "Стадия")
    ws.cell(1, 3, "Игрок 1")
    ws.cell(1, 4, "Игрок 2")
    ws.cell(1, 5, "Счёт по сетам")

    row = 2

    try:
        # Открываем страницу
        driver.get(tournament_url)

        # Небольшая пауза для загрузки данных
        time.sleep(5)

        # Находим элементы матчей
        matches = driver.find_elements(By.CSS_SELECTOR, "div.event__match")

        # Обрабатываем каждый матч
        for match in matches:
            try:
                # Получаем название турнира
                try:
                    tournament_name = driver.find_element(By.CSS_SELECTOR, "div.tournamentHeader__country").text
                except Exception:
                    tournament_name = "Неизвестный турнир"

                # Получаем стадию матча
                try:
                    stage = match.find_element(By.CSS_SELECTOR, "div.event__round").text
                except Exception:
                    stage = "Неизвестная стадия"

                # Получаем игроков
                try:
                    home_player = match.find_element(By.CSS_SELECTOR, "div.event__participant--home").text
                except Exception:
                    home_player = "Неизвестный игрок"

                try:
                    away_player = match.find_element(By.CSS_SELECTOR, "div.event__participant--away").text
                except Exception:
                    away_player = "Неизвестный игрок"

                # Получаем счёт по сетам
                try:
                    sets = match.find_elements(By.CSS_SELECTOR, "div.event__part")
                    set_scores = [set_elem.text for set_elem in sets]

                    # Форматирование счёта
                    formatted_scores = []
                    for set_score in set_scores:
                        # Разделяем по дефису, чтобы получить количество геймов
                        games = set_score.split('-')
                        if len(games) == 2:
                            # Если сет "7-6", то добавляем дополнительные цифры в скобках
                            if games[0] == '7' and games[1] == '6':
                                # Пример: 7(6) - 6(4)
                                score_home = games[0] + f"({sets[set_scores.index(set_score) + 1].text})"
                                score_away = games[1] + f"({sets[set_scores.index(set_score) + 2].text})"
                                formatted_scores.append(f"{score_home} - {score_away}")
                            else:
                                # Форматируем как X - Y
                                formatted_scores.append(f"{games[0]} - {games[1]}")
                        else:
                            formatted_scores.append(set_score)

                    score = " - ".join(formatted_scores)
                except Exception:
                    score = "Счёт недоступен"

                # Записываем данные в Excel
                ws.cell(row, 1, tournament_name)
                ws.cell(row, 2, stage)
                ws.cell(row, 3, home_player)
                ws.cell(row, 4, away_player)
                ws.cell(row, 5, score)

                row += 1

            except Exception as e:
                print(f"Ошибка обработки матча: {e}")

    finally:
        # Закрываем браузер
        driver.quit()

        # Автоподгонка ширины столбцов
        for col in range(1, 6):
            column = get_column_letter(col)
            max_length = max(len(str(ws[f"{column}{r}"].value)) for r in range(1, ws.max_row + 1))
            ws.column_dimensions[column].width = max_length + 2

        # Сохраняем файл Excel
        wb.save(file_path)
        print(f"Данные о матчах сохранены в Excel файл по пути: {file_path}")


# Основная функция
def main():
    tournament_url = "https://www.livesport.com/tennis/atp-singles/adelaide/results/"
    extract_matches_with_selenium(tournament_url)


if __name__ == "__main__":
    main()
