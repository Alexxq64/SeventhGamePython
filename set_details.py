from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import os
import time


def extract_point_by_point_links(tournament_url):
    """
    Извлекает ссылки на все вкладки "point-by-point" для каждого матча на указанной странице турнира.

    Параметры:
        tournament_url (str): URL страницы турнира.
    """
    # Настраиваем Selenium
    chrome_options = Options()
    chrome_options.add_argument("--headless")  # Открытие в фоновом режиме
    chrome_service = Service("C:\\Users\\User\\Desktop\\Настройки\\chromedriver.exe")

    # Инициализация браузера
    driver = webdriver.Chrome(service=chrome_service, options=chrome_options)

    # Путь к Excel файлу
    file_path = "D:\\Hobby\\SeventhGamePython\\7thGame.xlsx"

    try:
        # Открываем страницу турнира
        driver.get(tournament_url)

        # Небольшая пауза для загрузки данных
        time.sleep(5)

        # Находим элементы матчей
        matches = driver.find_elements(By.CSS_SELECTOR, "div.event__match")

        # Загружаем файл Excel или создаём новый
        wb = load_workbook(file_path) if os.path.exists(file_path) else Workbook()
        ws_links = wb.create_sheet("SetLinks") if "SetLinks" not in wb.sheetnames else wb["SetLinks"]
        ws_errors = wb.create_sheet("Errors") if "Errors" not in wb.sheetnames else wb["Errors"]

        # Заголовки для ссылок и ошибок
        ws_links.cell(1, 1, "Ссылка на матч")
        ws_links.cell(1, 2, "Ссылка на вкладку")
        ws_errors.cell(1, 1, "Ссылка на матч")
        ws_errors.cell(1, 2, "Ошибка")

        link_row = 2
        error_row = 2

        # Обрабатываем каждый элемент матча
        for match in matches:
            try:
                # Находим ссылку на матч
                link_element = match.find_element(By.CSS_SELECTOR, "a.eventRowLink")
                match_url = link_element.get_attribute("href")

                # Проверяем вкладки "point-by-point"
                for set_number in range(5):  # Проверяем вкладки от 0 до 4
                    point_by_point_url = f"{match_url}point-by-point/{set_number}"
                    driver.get(point_by_point_url)
                    time.sleep(2)  # Небольшая пауза для загрузки страницы

                    # Проверяем, существует ли вкладка
                    if "point-by-point" in driver.current_url:
                        # Записываем рабочую ссылку в Excel
                        ws_links.cell(link_row, 1, match_url)
                        ws_links.cell(link_row, 2, point_by_point_url)
                        link_row += 1
                    else:
                        break  # Если вкладка недоступна, выходим из цикла

            except Exception as e:
                # Если ошибка, записываем в лист ошибок
                ws_errors.cell(error_row, 1, match_url if 'match_url' in locals() else "Неизвестно")
                ws_errors.cell(error_row, 2, str(e))
                error_row += 1

        # Автоподгонка ширины столбцов
        for sheet in [ws_links, ws_errors]:
            for col in range(1, sheet.max_column + 1):
                column = get_column_letter(col)
                max_length = max(len(str(sheet[f'{column}{row}'].value)) for row in range(1, sheet.max_row + 1))
                sheet.column_dimensions[column].width = max_length + 2

        # Сохраняем файл Excel
        wb.save(file_path)

        # Выводим сообщение об успехе
        print(f"Ссылки на вкладки сохранены в Excel файл по пути: {file_path}")

    finally:
        # Закрываем браузер
        driver.quit()


# Основная функция
def main():
    tournament_url = "https://www.livesport.com/tennis/atp-singles/adelaide/results/"
    extract_point_by_point_links(tournament_url)


if __name__ == "__main__":
    main()
